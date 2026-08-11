import pandas as pd
import numpy as np
import datetime as dt
from pathlib import Path
from tkinter.filedialog import askopenfilename, asksaveasfilename
import shutil
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os

from CCI_modules.CCI_utils import VARNAME_MAPPING_NO, PATH_GENERERTE_FILER
from CCI_modules.CCI_utils import read_combined_historical_data, read_norway_historical_data
# LESE DATA

CURRENT_MONTH = dt.datetime.now().strftime('%Y-%m')

def get_output_path():
    output_path = PATH_GENERERTE_FILER.joinpath(CURRENT_MONTH)
    #
    if not(os.path.isdir(output_path)):
        os.mkdir(output_path)
        print('\nCreated folder: {}\n'.format(output_path))
    #
    return output_path


def create_cci_rapport_file():
    template_file = PATH_GENERERTE_FILER.joinpath('TEMPLAT/Opinion - CCI - TEMPLATE.xlsx')
    print('Template file: {}'.format(template_file))

    #today = dt.datetime.now().strftime('%y%m%d')
    #output_file = PATH_RAPPORTER.joinpath('Opinion - CCI - {}.xlsx'.format(today))
    output_file = 'Opinion - CCI - Abonnenter {}.xlsx'.format(CURRENT_MONTH)
    output_path = get_output_path()
    output_file = asksaveasfilename(initialdir = output_path, initialfile = output_file)

    # Kopier prev_excel_file til ny fil med dagens dato
    shutil.copyfile(template_file, output_file)
    print('\nCopied template file.\nCreated output file: {}\n'.format(output_file))
    input('Press Enter to continue')
    return output_file


def write_country_sheets_to_excel(data_samlet, output_file):
    # Skriv til Excel-fil
    wb = openpyxl.load_workbook(output_file, read_only=False)

    data = data_samlet.copy()
    #data.index = pd.to_datetime(data.index).strftime("%Y/%m")
    data.index = data.index.strftime("%Y/%m")

    # Loop through sheets
    countries = ['NO','DK','FI','SE','EA','CCI Nordic']

    for country in countries:
        print('********{}*********'.format(country))
        sheet=country
        col1 = pd.read_excel(output_file, sheet_name=sheet, usecols=[0], header=None)
        headerrow = np.where(col1=='year/month')[0][0]
        print(headerrow)   
        startrow=headerrow+1 
        sheet_data = pd.read_excel(output_file, sheet_name=sheet, header=headerrow, index_col=0)
        startdate = sheet_data.index[0] #sheet_data.loc[0,'year/month']
        #
        output_data = data.reindex(columns=sheet_data.columns)
        output_data = output_data.loc[output_data.index>=startdate]
        print(output_data)

        rows = dataframe_to_rows(output_data, index=True, header=False)
        ws=wb[sheet]
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                _ = ws.cell(row=r_idx+startrow, column=c_idx, value=value)

    print('\nSaving {} \n...\n\n'.format(output_file))
    wb.save(output_file)
    wb.close()
    print('File saved\n')



def write_norway_details_to_excel(norge_detalj, output_file):
    sheet = 'Norway details'
    print('********{}*********'.format(sheet))
    sheet_data = pd.read_excel(output_file, sheet_name=sheet, header=[5,6], index_col=0)
    col1 = pd.read_excel(output_file, sheet_name=sheet, usecols=[0], header=None)
    
    positive = ['Better', 'Higher', 'Probable', 'Net savings']
    negative = ['Worse', 'Lower', 'Not probable', 'Taking up a loan']
    level1_mapping = {p:'up' for p in positive}|{n:'down' for n in negative}|{'Net score': 'net'}
    level0_mapping = {v:k for k,v in VARNAME_MAPPING_NO.items()}

    temp = sheet_data.rename(columns=level1_mapping, level=1).rename(columns=level0_mapping, level=0)
    align_cols = temp.columns

    #Finn rad for angitt startdato
    start_from_date = '2024/01'
    startrow = np.where(col1==start_from_date)[0][0]-1

    # Tilpass output data - kolonner og rader
    output_data = norge_detalj.reindex(columns=align_cols)
    #output_data.index = pd.to_datetime(output_data.index, dayfirst=True).strftime("%Y/%m")
    output_data.index = output_data.index.strftime("%Y/%m")
    output_data = output_data.loc[output_data.index>=start_from_date]
    
    # Dele up/down på 100 for å skrive prosenter
    pct_cols = output_data.columns.get_level_values(1).isin(['up', 'down'])
    output_data.loc[:, pct_cols] = output_data.loc[:, pct_cols] / 100
    print(output_data)

    wb = openpyxl.load_workbook(output_file, read_only=False)
    ws=wb[sheet]
    rows = dataframe_to_rows(output_data, index=True, header=False)

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            _ = ws.cell(row=r_idx+startrow, column=c_idx, value=value)
    print('\nSaving {} \n...\n\n'.format(output_file))
    wb.save(output_file)
    wb.close()
    print('File saved\n')



def write_csv_norgesbank(data_norge_net):
    # Sletter tomme rader
    output_data = data_norge_net.dropna(how='all',axis=0).copy()
    # Tilpasser kolonnenavn
    output_data.columns = output_data.columns.str.replace(r'^(EU|NO)\d{2}_', '', regex=True)
    #output_data = output_data.reset_index().rename(columns={'index':'date'})
    output_data['date'] = output_data.index.strftime('%d/%m/%Y')
    #
    cols_to_write = ['date', 'CCI_Norge_gml', 'CCI', 'Kjopsindeksen',
        'Landets_oko_naa', 'Landets_oko_12mnd', 'Egen_oko_naa', 'Egen_oko_12mnd',
        'Arbeidsl_12mnd', 'Renter_12mnd', 'Hushold_finanser', 'Sparing_12mnd',
        'Store_kjop_12mnd', 'Bilkjop_12mnd', 'Boligkjop_12mnd', 'Oppussing_12mnd'
    ]
    output_data = output_data[cols_to_write]
    #
    output_file = 'Forbrukermeteret_Opinion_NorgesBank_{}.csv'.format(CURRENT_MONTH)
    output_path = get_output_path()
    output_file = asksaveasfilename(initialdir = output_path, initialfile = output_file)
    output_data.to_csv(output_file, sep=';', decimal=',', index=False)

    return output_data



def write_cci_reports(data_samlet, data_norge_detalj):
    print('Skriver Excel-rapport til abonnenter Forbrukermeteret')
    output_file = create_cci_rapport_file()
    write_country_sheets_to_excel(data_samlet, output_file)
    write_norway_details_to_excel(data_norge_detalj, output_file)

    print('Skriver fil til Norges Bank')
    data_norge_net = data_norge_detalj.xs('net', level=1, axis=1)
    _ = write_csv_norgesbank(data_norge_net)



def cci_excelrapport_main():
    data_samlet = read_combined_historical_data()
    data_norge_detalj = read_norway_historical_data()
    write_cci_reports(data_samlet, data_norge_detalj)

